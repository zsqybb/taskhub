"""数据导出蓝图 — Excel .xlsx"""
from io import BytesIO
from flask import Blueprint, send_file, request, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import query
from decorators import login_required

export_bp = Blueprint("export", __name__)

HEADER_FILL = PatternFill("solid", fgColor="6366F1")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _style_data(ws, start_row):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


STATUS_MAP = {
    "planning": "规划中", "in_progress": "进行中",
    "completed": "已完成", "cancelled": "已取消",
    "todo": "待办", "review": "评审中", "done": "已完成",
    "pending": "待开始",
}


@export_bp.get("/api/export/projects")
@login_required
def export_projects():
    rows = query("""
        SELECT p.name, p.description, p.status, p.priority, p.budget, p.progress,
               p.start_date, p.end_date, tm.name AS manager_name,
               COUNT(t.id) AS task_count,
               SUM(CASE WHEN t.status='done' THEN 1 ELSE 0 END) AS done_count
        FROM project p
        LEFT JOIN team_member tm ON p.manager_id = tm.id
        LEFT JOIN task t ON t.project_id = p.id
        WHERE p.owner_id = %s
        GROUP BY p.id ORDER BY p.created_at DESC
    """, [session["user_id"]])

    wb = Workbook()
    ws = wb.active
    ws.title = "项目列表"
    headers = ["项目名称", "描述", "状态", "优先级", "预算(元)", "进度(%)",
               "开始日期", "截止日期", "负责人", "任务数", "已完成"]
    ws.append(headers)
    _style_header(ws, 1)

    for r in rows:
        ws.append([
            r["name"], r.get("description"), STATUS_MAP.get(r["status"], r["status"]),
            {1: "低", 2: "中", 3: "高", 4: "紧急"}.get(r["priority"], r["priority"]),
            float(r["budget"] or 0), float(r["progress"] or 0),
            str(r.get("start_date") or ""), str(r.get("end_date") or ""),
            r.get("manager_name"), r.get("task_count"), r.get("done_count"),
        ])
    _style_data(ws, 2)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    for c in "CDEFGHIJK":
        ws.column_dimensions[c].width = 12

    return _send(wb, "projects.xlsx")


@export_bp.get("/api/export/tasks")
@login_required
def export_tasks():
    project_id = request.args.get("project_id")
    sql = """
        SELECT t.title, t.description, t.status, t.priority,
               t.estimated_hours, t.actual_hours, t.start_date, t.due_date,
               tm.name AS assignee_name, p.name AS project_name
        FROM task t
        LEFT JOIN team_member tm ON t.assignee_id = tm.id
        JOIN project p ON t.project_id = p.id
        WHERE p.owner_id = %s
    """
    args = [session["user_id"]]
    if project_id:
        sql += " AND t.project_id = %s"
        args.append(project_id)
    sql += " ORDER BY t.sort_order, t.id"
    rows = query(sql, args)

    wb = Workbook()
    ws = wb.active
    ws.title = "任务列表"
    headers = ["任务标题", "描述", "状态", "优先级", "预估工时(h)", "实际工时(h)",
               "开始日期", "截止日期", "负责人", "所属项目"]
    ws.append(headers)
    _style_header(ws, 1)

    for r in rows:
        ws.append([
            r["title"], r.get("description"),
            STATUS_MAP.get(r["status"], r["status"]),
            {1: "低", 2: "中", 3: "高", 4: "紧急"}.get(r["priority"], r["priority"]),
            float(r["estimated_hours"] or 0), float(r["actual_hours"] or 0),
            str(r.get("start_date") or ""), str(r.get("due_date") or ""),
            r.get("assignee_name"), r.get("project_name"),
        ])
    _style_data(ws, 2)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    for c in "CDEFGHIJ":
        ws.column_dimensions[c].width = 14

    return _send(wb, "tasks.xlsx")


@export_bp.get("/api/export/costs")
@login_required
def export_costs():
    project_id = request.args.get("project_id")
    sql = """
        SELECT c.cost_date, c.category, c.amount, c.description, p.name AS project_name
        FROM cost c JOIN project p ON c.project_id = p.id
        WHERE p.owner_id = %s
    """
    args = [session["user_id"]]
    if project_id:
        sql += " AND c.project_id = %s"
        args.append(project_id)
    sql += " ORDER BY c.cost_date DESC"
    rows = query(sql, args)

    wb = Workbook()
    ws = wb.active
    ws.title = "成本报表"
    headers = ["日期", "类别", "金额(元)", "说明", "所属项目"]
    ws.append(headers)
    _style_header(ws, 1)

    CAT_MAP = {"labor": "人力", "equipment": "设备", "software": "软件", "travel": "差旅", "other": "其他"}
    for r in rows:
        ws.append([
            str(r["cost_date"]), CAT_MAP.get(r["category"], r["category"]),
            float(r["amount"]), r.get("description"), r.get("project_name"),
        ])
    _style_data(ws, 2)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 20

    return _send(wb, "costs.xlsx")


def _send(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename,
    )
